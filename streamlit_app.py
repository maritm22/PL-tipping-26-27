
from __future__ import annotations
from pathlib import Path

import io
import re
from typing import Any

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="PL-Tippen 2026/27", page_icon="🏆", layout="wide")

FPL = "https://fantasy.premierleague.com/api"
SCORING = {0: 10, 1: 7, 2: 5, 3: 3}
BONUS_POINTS = 15

# ---------- Styling ----------
st.markdown("""
<style>
:root {
    --pl-purple: #37003c;
    --pl-green: #00ff87;
    --pl-cyan: #00c2ff;
    --pl-pink: #ff2882;
    --pl-gold: #f4c542;
    --pl-silver: #b8c2cc;
    --pl-bronze: #cd7f32;
}

.block-container {
    padding-top: 1.4rem;
    padding-bottom: 3rem;
}

h1, h2, h3 {
    letter-spacing: -0.02em;
}

[data-testid="stMetric"] {
    background: rgba(55, 0, 60, 0.045);
    border: 1px solid rgba(55, 0, 60, 0.12);
    padding: 10px 12px;
    border-radius: 14px;
}

.leader-card {
    border-radius: 18px;
    padding: 14px 18px;
    margin-bottom: 12px;
    border: 1px solid rgba(128,128,128,.22);
    background: linear-gradient(135deg, rgba(55,0,60,.06), rgba(0,194,255,.035));
}

.rank-1 {
    border-left: 7px solid var(--pl-gold);
    background: linear-gradient(135deg, rgba(244,197,66,.14), rgba(255,255,255,0));
}
.rank-2 {
    border-left: 7px solid var(--pl-silver);
    background: linear-gradient(135deg, rgba(184,194,204,.15), rgba(255,255,255,0));
}
.rank-3 {
    border-left: 7px solid var(--pl-bronze);
    background: linear-gradient(135deg, rgba(205,127,50,.14), rgba(255,255,255,0));
}
.rank-other {
    border-left: 7px solid var(--pl-purple);
}

.bonus-row {
    padding: 8px 11px;
    border-radius: 10px;
    margin: 6px 0;
    font-size: .92rem;
}
.bonus-ok {
    background: rgba(0, 200, 100, .14);
    border: 1px solid rgba(0, 170, 85, .35);
}
.bonus-no {
    background: rgba(220, 60, 70, .12);
    border: 1px solid rgba(220, 60, 70, .30);
}
.bonus-neutral {
    background: rgba(128,128,128,.09);
    border: 1px solid rgba(128,128,128,.22);
}

.section-title {
    background: linear-gradient(90deg, var(--pl-purple), #5d1466);
    color: white;
    padding: 10px 14px;
    border-radius: 12px;
    margin: 8px 0 14px 0;
    font-weight: 700;
}

.live-pill {
    display: inline-block;
    background: rgba(0,255,135,.14);
    color: inherit;
    border: 1px solid rgba(0,190,105,.38);
    border-radius: 999px;
    padding: 5px 10px;
    margin-bottom: 8px;
    font-weight: 600;
}

.small-muted {
    opacity: .72;
    font-size: .88rem;
}
</style>
""", unsafe_allow_html=True)

# ---------- API ----------
@st.cache_data(ttl=900, show_spinner=False)
def get_bootstrap() -> dict:
    r = requests.get(f"{FPL}/bootstrap-static/", timeout=20)
    r.raise_for_status()
    return r.json()

@st.cache_data(ttl=900, show_spinner=False)
def get_fixtures() -> list[dict]:
    r = requests.get(f"{FPL}/fixtures/", timeout=20)
    r.raise_for_status()
    return r.json()

@st.cache_data(ttl=3600, show_spinner=False)
def get_event_live(gw: int) -> dict:
    r = requests.get(f"{FPL}/event/{gw}/live/", timeout=20)
    r.raise_for_status()
    return r.json()

def refresh_live_data():
    st.cache_data.clear()
    st.rerun()

# ---------- Helpers ----------
def norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())

def norm_key(s: Any) -> str:
    return re.sub(r"[^a-z0-9æøå]+", " ", norm(s).lower()).strip()



def points_for_diff(diff: int) -> int:
    return SCORING.get(abs(int(diff)), 0)

def stats_from_diffs(diffs: list[int]) -> dict:
    d = [abs(int(x)) for x in diffs]
    n = len(d) or 1
    return {
        "exact": sum(x == 0 for x in d),
        "one": sum(x == 1 for x in d),
        "two": sum(x == 2 for x in d),
        "three": sum(x == 3 for x in d),
        "over": sum(x > 3 for x in d),
        "within1": 100 * sum(x <= 1 for x in d) / n,
        "avg": sum(d) / n,
    }

# ---------- Excel parsing ----------
BONUS_ALIASES = {
    "Flest mål": ["flest mål", "flest mal", "mest mål", "mest mal"],
    "Flest clean sheets": ["flest clean sheets", "clean sheets", "flest nuller", "flest holdt nullen"],
    "Flest gule kort": ["flest gule kort", "gule kort"],
    "Flest røde kort": ["flest røde kort", "flest rode kort", "røde kort", "rode kort"],
    "Flest uavgjorte": ["flest uavgjorte", "uavgjorte"],
    "Toppscorer": ["toppscorer", "toppskårer", "toppskarer"],
    "Flest assist": ["flest assist", "assist"],
}

def find_value_right_of_label(ws, aliases: list[str]):
    alias_keys = [norm_key(a) for a in aliases]
    for row in ws.iter_rows():
        for c in row:
            ck = norm_key(c.value)
            if not ck:
                continue
            if any(a in ck for a in alias_keys):
                # First useful value to the right (up to 5 columns)
                for j in range(c.column + 1, min(c.column + 6, ws.max_column + 1)):
                    v = ws.cell(c.row, j).value
                    if v not in (None, ""):
                        return norm(v)
    return ""

def parse_participant_xlsx(file_obj) -> tuple[dict, list[str]]:
    data = file_obj.getvalue() if hasattr(file_obj, "getvalue") else file_obj.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    warnings = []

    # Prefer known sheets, but stay tolerant
    tab_ws = wb["Tabelltips"] if "Tabelltips" in wb.sheetnames else wb[wb.sheetnames[0]]
    bonus_ws = wb["Bonustips"] if "Bonustips" in wb.sheetnames else (
        wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else tab_ws
    )

    # Participant name: find a dedicated name field, else infer from filename.
    participant = ""
    rejected_name_values = {
        "kategori", "tips", "lag", "plass", "plassering", "tabelltips",
        "bonustips", "spiller", "svar", "premier league"
    }

    for ws in (tab_ws, bonus_ws):
        for row in ws.iter_rows():
            for c in row:
                label = norm_key(c.value)
                if label in ("navn", "deltaker", "deltakernavn", "navn deltaker"):
                    for j in range(c.column + 1, min(c.column + 5, ws.max_column + 1)):
                        v = norm(ws.cell(c.row, j).value)
                        vk = norm_key(v)
                        if v and vk not in rejected_name_values:
                            participant = v
                            break
                if participant:
                    break
            if participant:
                break
        if participant:
            break

    filename = getattr(file_obj, "name", "")
    stem = Path(filename).stem
    inferred = re.sub(r"(?i)^premier league[- –—]*tips[- –—]*", "", stem).strip()
    inferred = re.sub(r"\(\d+\)$", "", inferred).strip()

    generic_names = {
        "kategori", "tips", "lag", "plass", "plassering", "tabelltips",
        "bonustips", "spiller", "svar", "svartype", "kategori svartype",
        "premier league"
    }

    # For this competition, the submitted filename is the safest participant identifier:
    # "Premier League-tips – Ola Nordmann.xlsx" -> "Ola Nordmann".
    filename_has_participant = bool(re.match(
        r"(?i)^premier league[- –—]*tips[- –—]*.+", stem
    ))
    parsed_participant = participant

    if filename_has_participant and inferred:
        participant = inferred
    elif (not participant) or norm_key(participant) in generic_names:
        participant = inferred or stem

    # Only warn when the sheet contains a plausible real name that disagrees with the filename.
    if (
        inferred and parsed_participant
        and norm_key(parsed_participant) not in generic_names
        and norm_key(inferred) not in norm_key(parsed_participant)
        and norm_key(parsed_participant) not in norm_key(inferred)
    ):
        warnings.append(
            f'Filnavnet tyder på "{inferred}", mens et navnefelt i arket ser ut til å være "{parsed_participant}".'
        )

    # Table tips: scan rows for team + integer rank 1..20
    tips = {}
    for row in tab_ws.iter_rows(values_only=True):
        vals = [v for v in row if v not in (None, "")]
        if len(vals) < 2:
            continue
        ints = []
        strings = []
        for v in vals:
            if isinstance(v, (int, float)) and int(v) == v and 1 <= int(v) <= 20:
                ints.append(int(v))
            elif isinstance(v, str):
                strings.append(norm(v))
        if ints and strings:
            # Pick the most team-looking string, avoiding headings
            candidates = [s for s in strings if norm_key(s) not in (
                "plass", "plassering", "lag", "tabelltips", "tips", "premier league"
            ) and len(s) > 2]
            if candidates:
                team = candidates[-1]
                tips[team] = ints[0]

    # If scan captured too much/little, keep only unique ranks where possible
    if len(tips) != 20:
        warnings.append(f"Fant {len(tips)} tabelltips; forventet 20.")

    ranks = list(tips.values())
    if len(ranks) != len(set(ranks)):
        warnings.append("Det finnes dupliserte plasseringer i tabelltipset.")
    if set(ranks) and set(ranks) != set(range(1, 21)):
        warnings.append("Plasseringene dekker ikke nøyaktig 1–20.")

    bonus = {}
    for label, aliases in BONUS_ALIASES.items():
        bonus[label] = find_value_right_of_label(bonus_ws, aliases)
        if not bonus[label]:
            warnings.append(f'Mangler svar på "{label}".')

    return {"name": participant, "tips": tips, "bonus": bonus, "filename": filename}, warnings

# ---------- Football model ----------
def fpl_maps(bootstrap: dict):
    team_by_id = {t["id"]: t["name"] for t in bootstrap["teams"]}
    player_by_id = {}
    for p in bootstrap["elements"]:
        player_by_id[p["id"]] = {
            "name": p["web_name"],
            "full_name": f'{p.get("first_name","")} {p.get("second_name","")}'.strip(),
            "team": team_by_id.get(p["team"], str(p["team"])),
            "team_id": p["team"],
            "goals": int(p.get("goals_scored", 0)),
            "assists": int(p.get("assists", 0)),
            "yellow": int(p.get("yellow_cards", 0)),
            "red": int(p.get("red_cards", 0)),
        }
    return team_by_id, player_by_id

def latest_finished_gw(fixtures: list[dict]) -> int:
    finished_events = [int(f["event"]) for f in fixtures if f.get("finished") and f.get("event")]
    return max(finished_events, default=0)

def standings_through_gw(bootstrap: dict, fixtures: list[dict], gw: int) -> pd.DataFrame:
    team_by_id, _ = fpl_maps(bootstrap)
    rows = {
        tid: {"Lag": name, "K": 0, "V": 0, "U": 0, "T": 0, "MF": 0, "MM": 0, "P": 0, "CS": 0}
        for tid, name in team_by_id.items()
    }

    for f in fixtures:
        ev = f.get("event")
        if not f.get("finished") or not ev or int(ev) > gw:
            continue
        h, a = f["team_h"], f["team_a"]
        hg, ag = int(f["team_h_score"]), int(f["team_a_score"])
        rows[h]["K"] += 1; rows[a]["K"] += 1
        rows[h]["MF"] += hg; rows[h]["MM"] += ag
        rows[a]["MF"] += ag; rows[a]["MM"] += hg
        if ag == 0: rows[h]["CS"] += 1
        if hg == 0: rows[a]["CS"] += 1
        if hg > ag:
            rows[h]["V"] += 1; rows[a]["T"] += 1; rows[h]["P"] += 3
        elif hg < ag:
            rows[a]["V"] += 1; rows[h]["T"] += 1; rows[a]["P"] += 3
        else:
            rows[h]["U"] += 1; rows[a]["U"] += 1
            rows[h]["P"] += 1; rows[a]["P"] += 1

    df = pd.DataFrame(rows.values())
    df["MF"] = df["MF"].astype(int)
    df["MM"] = df["MM"].astype(int)
    df["MF"] = df["MF"].fillna(0)
    df["MM"] = df["MM"].fillna(0)
    df["MD"] = df["MF"] - df["MM"]

    if gw == 0 or df["K"].sum() == 0:
        df = df.sort_values(["Lag"], ascending=[True], kind="stable")
    else:
        df = df.sort_values(["P", "MD", "MF", "Lag"], ascending=[False, False, False, True], kind="stable")
    df = df.reset_index(drop=True)
    df.insert(0, "Pl", range(1, len(df) + 1))
    return df[["Pl", "Lag", "K", "V", "U", "T", "MF", "MM", "MD", "P", "CS"]]

def current_player_totals(bootstrap: dict) -> pd.DataFrame:
    team_by_id, _ = fpl_maps(bootstrap)
    data = []
    for p in bootstrap["elements"]:
        data.append({
            "id": p["id"],
            "Spiller": p["web_name"],
            "Lag": team_by_id.get(p["team"], ""),
            "Mål": int(p.get("goals_scored", 0)),
            "Assist": int(p.get("assists", 0)),
            "Gule": int(p.get("yellow_cards", 0)),
            "Røde": int(p.get("red_cards", 0)),
        })
    return pd.DataFrame(data)

def cumulative_player_totals_through_gw(bootstrap: dict, gw: int) -> pd.DataFrame:
    team_by_id, player_by_id = fpl_maps(bootstrap)
    totals = {
        pid: {"id": pid, "Spiller": p["name"], "Lag": p["team"], "Mål": 0, "Assist": 0, "Gule": 0, "Røde": 0}
        for pid, p in player_by_id.items()
    }
    for g in range(1, gw + 1):
        try:
            live = get_event_live(g)
        except Exception:
            continue
        for e in live.get("elements", []):
            pid = int(e["id"])
            if pid not in totals:
                continue
            s = e.get("stats", {})
            totals[pid]["Mål"] += int(s.get("goals_scored", 0))
            totals[pid]["Assist"] += int(s.get("assists", 0))
            totals[pid]["Gule"] += int(s.get("yellow_cards", 0))
            totals[pid]["Røde"] += int(s.get("red_cards", 0))
    return pd.DataFrame(totals.values())

def leaders_from_df(df: pd.DataFrame, value_col: str, name_col: str) -> tuple[list[str], int]:
    if df.empty:
        return [], 0
    maxv = int(df[value_col].max())
    if maxv <= 0:
        return [], 0
    return sorted(df.loc[df[value_col] == maxv, name_col].dropna().astype(str).unique().tolist()), maxv

def bonus_state(bootstrap: dict, fixtures: list[dict], gw: int, historical_players: pd.DataFrame | None = None):
    table = standings_through_gw(bootstrap, fixtures, gw)
    if gw == 0:
        return {
            k: {"leaders": [], "value": 0, "unit": ""}
            for k in BONUS_ALIASES
        }

    players = historical_players if historical_players is not None else current_player_totals(bootstrap)

    # Team cards from player totals
    team_cards = players.groupby("Lag", as_index=False)[["Gule", "Røde"]].sum()

    top_goals, v_goals = leaders_from_df(table, "MF", "Lag")
    top_cs, v_cs = leaders_from_df(table, "CS", "Lag")
    top_draws, v_draws = leaders_from_df(table, "U", "Lag")
    top_yellow, v_yellow = leaders_from_df(team_cards, "Gule", "Lag")
    top_red, v_red = leaders_from_df(team_cards, "Røde", "Lag")
    top_scorer, v_scorer = leaders_from_df(players, "Mål", "Spiller")
    top_assist, v_assist = leaders_from_df(players, "Assist", "Spiller")

    return {
        "Flest mål": {"leaders": top_goals, "value": v_goals, "unit": "mål"},
        "Flest clean sheets": {"leaders": top_cs, "value": v_cs, "unit": "CS"},
        "Flest gule kort": {"leaders": top_yellow, "value": v_yellow, "unit": "gule"},
        "Flest røde kort": {"leaders": top_red, "value": v_red, "unit": "røde"},
        "Flest uavgjorte": {"leaders": top_draws, "value": v_draws, "unit": "uavgj."},
        "Toppscorer": {"leaders": top_scorer, "value": v_scorer, "unit": "mål"},
        "Flest assist": {"leaders": top_assist, "value": v_assist, "unit": "assist"},
    }

def answer_matches(answer: str, leaders: list[str]) -> bool:
    a = norm_key(answer)
    if not a:
        return False
    for x in leaders:
        xk = norm_key(x)
        if a == xk or a in xk or xk in a:
            return True
    return False

def align_tips_to_table(tips: dict[str, int], table: pd.DataFrame):
    current_pos = {norm_key(r["Lag"]): int(r["Pl"]) for _, r in table.iterrows()}
    diffs = []
    detail = []
    unmatched = []
    for team, predicted in tips.items():
        tk = norm_key(team)
        # exact, then forgiving contains
        actual = current_pos.get(tk)
        actual_team = team
        if actual is None:
            matches = [(k, p) for k, p in current_pos.items() if tk in k or k in tk]
            if len(matches) == 1:
                _, actual = matches[0]
        if actual is None:
            unmatched.append(team)
            continue
        diff = abs(int(predicted) - int(actual))
        diffs.append(diff)
        detail.append((team, int(predicted), int(actual), diff, points_for_diff(diff)))
    return diffs, detail, unmatched

def score_participant(participant: dict, table: pd.DataFrame, bonus: dict):
    diffs, detail, unmatched = align_tips_to_table(participant["tips"], table)
    table_points = sum(x[4] for x in detail)
    bonus_hits = {}
    bonus_points = 0
    for label, answer in participant["bonus"].items():
        leaders = bonus.get(label, {}).get("leaders", [])
        hit = bool(leaders) and answer_matches(answer, leaders)
        bonus_hits[label] = hit if leaders else None
        if hit:
            bonus_points += BONUS_POINTS
    return {
        "name": participant["name"],
        "participant": participant,
        "diffs": diffs,
        "detail": detail,
        "unmatched": unmatched,
        "table_points": table_points,
        "bonus_points": bonus_points,
        "total": table_points + bonus_points,
        "bonus_hits": bonus_hits,
        "stats": stats_from_diffs(diffs),
    }

# ---------- UI components ----------
def render_bonus_rows(scored: dict, bonus_state_now: dict):
    for label in BONUS_ALIASES:
        answer = scored["participant"]["bonus"].get(label, "") or "—"
        status = scored["bonus_hits"].get(label)
        if status is True:
            cls, icon = "bonus-ok", "🟢"
        elif status is False:
            cls, icon = "bonus-no", "🔴"
        else:
            cls, icon = "bonus-neutral", "⚪"
        st.markdown(
            f'<div class="bonus-row {cls}"><b>{icon} {label}</b><br>{answer}</div>',
            unsafe_allow_html=True,
        )

def render_participant_card(scored: dict, rank: int, bonus_state_now: dict):
    s = scored["stats"]
    rank_class = {1: "rank-1", 2: "rank-2", 3: "rank-3"}.get(rank, "rank-other")
    medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(rank, f"#{rank}")

    st.markdown(f'<div class="leader-card {rank_class}">', unsafe_allow_html=True)
    left, right = st.columns([1.55, 1], gap="large")

    with left:
        h1, h2, h3 = st.columns([2.2, 1, 1])
        with h1:
            st.subheader(f"{medal} {scored['name']}")
        with h2:
            st.metric("Total", scored["total"])
        with h3:
            st.metric("Bonus", scored["bonus_points"])

        st.caption(f"Tabellpoeng: {scored['table_points']} · Bonuspoeng: {scored['bonus_points']}")

        a,b,c,d,e = st.columns(5)
        a.metric("🎯 Fulltreff", s["exact"])
        b.metric("±1", s["one"])
        c.metric("±2", s["two"])
        d.metric("±3", s["three"])
        e.metric(">3", s["over"])

        x,y = st.columns(2)
        x.metric("Treffsikkerhet ±1", f'{s["within1"]:.0f}%')
        y.metric("Snittavvik", f'{s["avg"]:.1f} pl.')

    with right:
        st.markdown("#### Bonus")
        render_bonus_rows(scored, bonus_state_now)

    st.markdown("</div>", unsafe_allow_html=True)

def show_real_table(table: pd.DataFrame):
    d = table.rename(columns={"Pl":"#","K":"K","V":"V","U":"U","T":"T","MF":"MF","MM":"MM","MD":"MD","P":"P"})
    st.dataframe(
        d[["#","Lag","K","V","U","T","MF","MM","MD","P"]],
        hide_index=True,
        use_container_width=True,
        column_config={
            "#": st.column_config.NumberColumn(width="small"),
            "Lag": st.column_config.TextColumn(width="medium"),
        },
    )

def show_bonus_dashboard(bonus: dict):
    cols = st.columns(2)
    for i, label in enumerate(BONUS_ALIASES):
        info = bonus[label]
        leaders = ", ".join(info["leaders"]) if info["leaders"] else "Ikke i gang"
        val = f'{info["value"]} {info["unit"]}'.strip() if info["leaders"] else "—"
        with cols[i % 2]:
            with st.container(border=True):
                st.caption(label)
                st.subheader(leaders)
                st.write(val)

def build_history(participants: list[dict], bootstrap: dict, fixtures: list[dict], current_gw: int) -> pd.DataFrame:
    records = []
    cumulative_players = None
    for gw in range(0, current_gw + 1):
        table = standings_through_gw(bootstrap, fixtures, gw)
        if gw == 0:
            bstate = bonus_state(bootstrap, fixtures, 0)
        else:
            cumulative_players = cumulative_player_totals_through_gw(bootstrap, gw)
            bstate = bonus_state(bootstrap, fixtures, gw, cumulative_players)
        for p in participants:
            s = score_participant(p, table, bstate)
            records.append({"GW": gw, "Deltaker": p["name"], "Poeng": s["total"]})
    return pd.DataFrame(records)

# ---------- App ----------
st.title("🏆 PL-Tippen 2026/27")
st.caption("Live dashboard · tabelltips · bonuskonkurranse · utvikling gjennom sesongen")
st.markdown('<div class="live-pill">● LIVE SESONGDASHBOARD</div>', unsafe_allow_html=True)

with st.sidebar:
    st.header("Konkurranse")
    uploaded = st.file_uploader(
        "Last opp tippeark",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Du kan laste opp alle deltakernes Excel-filer samtidig.",
    )
    if st.button("🔄 Oppdater live-data", use_container_width=True):
        refresh_live_data()
    st.caption("Fotballdata caches i opptil 15 minutter mellom oppdateringer.")

if not uploaded:
    st.info("Last opp minst ett tippeark i sidepanelet for å starte.")
    st.stop()

participants = []
all_warnings = []
for f in uploaded:
    try:
        p, warnings = parse_participant_xlsx(f)
        participants.append(p)
        for w in warnings:
            all_warnings.append(f"**{f.name}:** {w}")
    except Exception as e:
        all_warnings.append(f"**{f.name}:** Kunne ikke lese filen ({e}).")

if all_warnings:
    with st.expander(f"⚠️ Kontroll av innleveringer ({len(all_warnings)} merknader)"):
        for w in all_warnings:
            st.markdown(f"- {w}")

if not participants:
    st.error("Fant ingen lesbare tippeark.")
    st.stop()

try:
    bootstrap = get_bootstrap()
    fixtures = get_fixtures()
    current_gw = latest_finished_gw(fixtures)
    table = standings_through_gw(bootstrap, fixtures, current_gw)
    current_bonus = bonus_state(bootstrap, fixtures, current_gw)
except Exception as e:
    st.error(f"Kunne ikke hente FPL-data akkurat nå: {e}")
    st.stop()

gw_label = f"GW{current_gw}" if current_gw else "GW0 – alfabetisk starttabell"
st.markdown(f"### Status: {gw_label}")

scored = [score_participant(p, table, current_bonus) for p in participants]
scored = sorted(scored, key=lambda x: (-x["total"], x["name"].lower()))

# Tabs
tab_dash, tab_table, tab_bonus, tab_history = st.tabs(
    ["🏆 Dashboard", "📋 Premier League-tabell", "🎯 Bonusstatistikk", "📈 Poengutvikling"]
)

with tab_dash:
    # Compact leaderboard
    leaderboard = pd.DataFrame([
        {"#": i+1, "Deltaker": s["name"], "Poeng": s["total"],
         "Tabell": s["table_points"], "Bonus": s["bonus_points"]}
        for i, s in enumerate(scored)
    ])
    st.dataframe(leaderboard, hide_index=True, use_container_width=True)

    st.markdown('<div class="section-title">Deltakerkort</div>', unsafe_allow_html=True)
    for rank, s in enumerate(scored, 1):
        render_participant_card(s, rank, current_bonus)

with tab_table:
    st.markdown(f'<div class="section-title">Premier League-tabell · {gw_label}</div>', unsafe_allow_html=True)
    if current_gw == 0:
        st.caption("Ingen kamper er ferdigspilt ennå. Alle står med 0 poeng og rangeres alfabetisk.")
    show_real_table(table)
    st.caption("Tabellen beregnes fra ferdigspilte FPL-registrerte Premier League-kamper.")

with tab_bonus:
    st.markdown(f'<div class="section-title">Bonusstatistikk · {gw_label}</div>', unsafe_allow_html=True)
    if current_gw == 0:
        st.info("Bonuskonkurransene starter først når det finnes sportslige resultater.")
    show_bonus_dashboard(current_bonus)

    if current_gw > 0:
        st.markdown("### Detaljer")
        players = current_player_totals(bootstrap)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Toppscorere**")
            st.dataframe(players.sort_values(["Mål","Spiller"], ascending=[False,True])[["Spiller","Lag","Mål"]].head(10),
                         hide_index=True, use_container_width=True)
        with c2:
            st.markdown("**Flest assist**")
            st.dataframe(players.sort_values(["Assist","Spiller"], ascending=[False,True])[["Spiller","Lag","Assist"]].head(10),
                         hide_index=True, use_container_width=True)

        team_cards = players.groupby("Lag", as_index=False)[["Gule","Røde"]].sum()
        c3, c4 = st.columns(2)
        with c3:
            st.markdown("**Gule kort per lag**")
            st.dataframe(team_cards.sort_values(["Gule","Lag"], ascending=[False,True]).head(20),
                         hide_index=True, use_container_width=True)
        with c4:
            st.markdown("**Røde kort per lag**")
            st.dataframe(team_cards.sort_values(["Røde","Lag"], ascending=[False,True]).head(20),
                         hide_index=True, use_container_width=True)

with tab_history:
    st.markdown('<div class="section-title">Poengutvikling gjennom sesongen</div>', unsafe_allow_html=True)
    st.caption("X-akse = Gameweek · Y-akse = live poeng dersom sesongen hadde sluttet etter den runden.")
    if current_gw == 0:
        hist = pd.DataFrame([{"GW":0, "Deltaker":s["name"], "Poeng":s["total"]} for s in scored])
    else:
        with st.spinner("Bygger GW-historikk..."):
            hist = build_history(participants, bootstrap, fixtures, current_gw)

    # pivot_table is defensive against accidental duplicate GW/deltaker rows.
    wide = hist.pivot_table(
        index="GW",
        columns="Deltaker",
        values="Poeng",
        aggfunc="max"
    ).sort_index()
    st.line_chart(wide, x_label="Gameweek", y_label="Poeng", height=480)

    st.markdown("#### Poeng per GW")
    st.dataframe(wide.reset_index(), hide_index=True, use_container_width=True)
